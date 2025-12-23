"""
Data Export Features
Export agent executions, HITL records, and analytics data
"""

import logging
import csv
import json
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
from io import StringIO, BytesIO
from enum import Enum

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Supported export formats"""
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    JSONL = "jsonl"


class DataExporter:
    """
    Data export manager
    Handles exporting various data types to multiple formats
    """
    
    def __init__(self):
        self.supported_formats = [ExportFormat.CSV, ExportFormat.JSON, ExportFormat.JSONL]
        
        if OPENPYXL_AVAILABLE:
            self.supported_formats.append(ExportFormat.EXCEL)
        
        logger.info(f"Data exporter initialized with formats: {self.supported_formats}")
    
    def export_executions(
        self,
        executions: List[Dict[str, Any]],
        format: ExportFormat = ExportFormat.CSV,
        include_metadata: bool = True
    ) -> bytes:
        """
        Export agent execution logs
        
        Args:
            executions: List of execution records
            format: Export format
            include_metadata: Include additional metadata
            
        Returns:
            Exported data as bytes
        """
        logger.info(f"Exporting {len(executions)} executions to {format}")
        
        if format == ExportFormat.CSV:
            return self._export_to_csv(executions, "executions")
        elif format == ExportFormat.JSON:
            return self._export_to_json(executions)
        elif format == ExportFormat.JSONL:
            return self._export_to_jsonl(executions)
        elif format == ExportFormat.EXCEL:
            return self._export_to_excel(executions, "Executions")
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def export_hitl_records(
        self,
        records: List[Dict[str, Any]],
        format: ExportFormat = ExportFormat.CSV,
        include_decisions: bool = True
    ) -> bytes:
        """
        Export HITL records
        
        Args:
            records: List of HITL records
            format: Export format
            include_decisions: Include decision details
            
        Returns:
            Exported data as bytes
        """
        logger.info(f"Exporting {len(records)} HITL records to {format}")
        
        if format == ExportFormat.CSV:
            return self._export_to_csv(records, "hitl_records")
        elif format == ExportFormat.JSON:
            return self._export_to_json(records)
        elif format == ExportFormat.JSONL:
            return self._export_to_jsonl(records)
        elif format == ExportFormat.EXCEL:
            return self._export_to_excel(records, "HITL Records")
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def export_analytics(
        self,
        analytics_data: Dict[str, Any],
        format: ExportFormat = ExportFormat.JSON
    ) -> bytes:
        """
        Export analytics and metrics
        
        Args:
            analytics_data: Analytics data dictionary
            format: Export format
            
        Returns:
            Exported data as bytes
        """
        logger.info(f"Exporting analytics to {format}")
        
        if format == ExportFormat.JSON:
            return self._export_to_json(analytics_data)
        elif format == ExportFormat.EXCEL:
            # Convert analytics dict to list format for Excel
            data_list = self._flatten_analytics(analytics_data)
            return self._export_to_excel(data_list, "Analytics")
        else:
            raise ValueError(f"Analytics export only supports JSON and EXCEL formats")
    
    def _export_to_csv(self, data: List[Dict[str, Any]], filename: str) -> bytes:
        """Export data to CSV format"""
        if not data:
            return b""
        
        output = StringIO()
        
        # Get all unique keys
        all_keys = set()
        for item in data:
            all_keys.update(self._flatten_dict(item).keys())
        
        fieldnames = sorted(all_keys)
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for item in data:
            flat_item = self._flatten_dict(item)
            writer.writerow(flat_item)
        
        return output.getvalue().encode('utf-8')
    
    def _export_to_json(self, data: Any) -> bytes:
        """Export data to JSON format"""
        return json.dumps(data, indent=2, default=str).encode('utf-8')
    
    # def _export_to_jsonl(self, data: List[Dict[str, Any]]) -> bytes:
    #     """Export data to JSON Lines format"""
    #     output = StringIO()
    #     for item in data:
    #         output.write(json.dumps(item, default=str) + '\n')
    #     return output.getvalue().encode('utf-8')

    def _export_to_jsonl(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data to JSON Lines format"""
        lines = []
        for item in data:
            lines.append(json.dumps(item, default=str))
        
        # Join with newlines and encode to bytes
        output_str = '\n'.join(lines)
        if output_str:
            output_str += '\n'
        
        return output_str.encode('utf-8')
    
    def _export_to_excel(self, data: List[Dict[str, Any]], sheet_name: str) -> bytes:
        """Export data to Excel format"""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed")
        
        if not data:
            # Return empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Get all unique keys
        all_keys = set()
        for item in data:
            all_keys.update(self._flatten_dict(item).keys())
        
        headers = sorted(all_keys)
        
        # Write headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, item in enumerate(data, start=2):
            flat_item = self._flatten_dict(item)
            for col_idx, header in enumerate(headers, start=1):
                value = flat_item.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert lists to comma-separated strings
                items.append((new_key, ', '.join(str(x) for x in v)))
            else:
                items.append((new_key, v))
        
        return dict(items)
    
    def _flatten_analytics(self, analytics: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Convert analytics dict to list of dicts for Excel export"""
        result = []
        
        for category, data in analytics.items():
            if isinstance(data, dict):
                for key, value in data.items():
                    result.append({
                        "category": category,
                        "metric": key,
                        "value": value
                    })
            else:
                result.append({
                    "category": category,
                    "metric": "value",
                    "value": data
                })
        
        return result
    def _compute_analytics(self, executions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Compute analytics from execution data"""
        if not executions:
            return {
                "total_executions": 0,
                "successful_executions": 0,
                "failed_executions": 0,
                "success_rate": 0,
                "failure_rate": 0,
                "avg_duration_ms": 0,
                "min_duration_ms": 0,
                "max_duration_ms": 0
            }
        
        total = len(executions)
        successful = sum(1 for e in executions if e.get("status") == "completed")
        failed = sum(1 for e in executions if e.get("status") == "failed")
        
        durations = [e.get("duration_ms", 0) for e in executions if e.get("duration_ms")]
        avg_duration = sum(durations) / len(durations) if durations else 0
        
        return {
            "total_executions": total,
            "successful_executions": successful,
            "failed_executions": failed,
            "success_rate": round((successful / total * 100), 2) if total > 0 else 0,
            "failure_rate": round((failed / total * 100), 2) if total > 0 else 0,
            "avg_duration_ms": round(avg_duration, 2),
            "min_duration_ms": min(durations) if durations else 0,
            "max_duration_ms": max(durations) if durations else 0
        }

class ReportGenerator:
    """
    Generate analytical reports
    Combines data export with analytics
    """
    
    def __init__(self, exporter: DataExporter):
        self.exporter = exporter
    
    def generate_execution_report(
        self,
        executions: List[Dict[str, Any]],
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        format: ExportFormat = ExportFormat.EXCEL
    ) -> bytes:
        """
        Generate comprehensive execution report
        
        Args:
            executions: List of execution records
            start_date: Filter start date
            end_date: Filter end date
            format: Export format
            
        Returns:
            Report as bytes
        """
        from datetime import timezone as tz
        # Filter by date if provided
        if start_date and start_date.tzinfo is None:
            start_date = start_date.replace(tzinfo=tz.utc)
        
        if end_date and end_date.tzinfo is None:
            end_date = end_date.replace(tzinfo=tz.utc)
        
        # Calculate analytics
        analytics = self._calculate_execution_analytics(executions)
        
        if format == ExportFormat.EXCEL and OPENPYXL_AVAILABLE:
            return self._generate_excel_report(executions, analytics)
        else:
            # For non-Excel formats, just export the executions
            return self.exporter.export_executions(executions, format)
    
    def generate_hitl_report(
        self,
        records: List[Dict[str, Any]],
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        format: ExportFormat = ExportFormat.EXCEL
    ) -> bytes:
        """
        Generate HITL analysis report
        
        Args:
            records: List of HITL records
            start_date: Filter start date
            end_date: Filter end date
            format: Export format
            
        Returns:
            Report as bytes
        """
        # Filter by date if provided
        if start_date or end_date:
            records = self._filter_by_date(records, start_date, end_date)
        
        # Calculate analytics
        analytics = self._calculate_hitl_analytics(records)
        
        if format == ExportFormat.EXCEL and OPENPYXL_AVAILABLE:
            return self._generate_hitl_excel_report(records, analytics)
        else:
            return self.exporter.export_hitl_records(records, format)
    
    def _filter_by_date(
        self,
        records: List[Dict[str, Any]],
        start_date: Optional[datetime],
        end_date: Optional[datetime]
    ) -> List[Dict[str, Any]]:
        """Filter records by date range"""
        filtered = records
        
        if start_date:
            filtered = [
                r for r in filtered
                if self._parse_date(r.get('created_at')) >= start_date
            ]
        
        if end_date:
            filtered = [
                r for r in filtered
                if self._parse_date(r.get('created_at')) <= end_date
            ]
        
        return filtered
    
    def _parse_date(self, date_str: Any) -> datetime:
        """Parse date string to datetime"""
        if isinstance(date_str, datetime):
            return date_str
        if isinstance(date_str, str):
            try:
                return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            except:
                return datetime.min
        return datetime.min
    
    def _calculate_execution_analytics(self, executions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate execution analytics"""
        if not executions:
            return {}
        
        total = len(executions)
        success_count = sum(1 for e in executions if e.get('status') == 'completed')
        failed_count = sum(1 for e in executions if e.get('status') == 'failed')
        
        durations = [e.get('duration_ms', 0) for e in executions if e.get('duration_ms')]
        
        return {
            "total_executions": total,
            "successful": success_count,
            "failed": failed_count,
            "success_rate": round(success_count / total * 100, 2) if total > 0 else 0,
            "avg_duration_ms": round(sum(durations) / len(durations), 2) if durations else 0,
            "min_duration_ms": min(durations) if durations else 0,
            "max_duration_ms": max(durations) if durations else 0
        }
    
    def _calculate_hitl_analytics(self, records: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate HITL analytics"""
        if not records:
            return {}
        
        total = len(records)
        approved = sum(1 for r in records if r.get('status') == 'approved')
        rejected = sum(1 for r in records if r.get('status') == 'rejected')
        pending = sum(1 for r in records if r.get('status') == 'pending')
        
        return {
            "total_records": total,
            "approved": approved,
            "rejected": rejected,
            "pending": pending,
            "approval_rate": round(approved / total * 100, 2) if total > 0 else 0,
            "rejection_rate": round(rejected / total * 100, 2) if total > 0 else 0
        }
    
    def _generate_excel_report(
        self,
        executions: List[Dict[str, Any]],
        analytics: Dict[str, Any]
    ) -> bytes:
        """Generate Excel report with multiple sheets"""
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        row = 1
        for key, value in analytics.items():
            ws_summary.cell(row=row, column=1, value=key)
            ws_summary.cell(row=row, column=2, value=value)
            row += 1
        
        # Executions sheet
        ws_executions = wb.create_sheet("Executions")
        
        if executions:
            headers = list(self.exporter._flatten_dict(executions[0]).keys())
            for col_idx, header in enumerate(headers, start=1):
                ws_executions.cell(row=1, column=col_idx, value=header)
            
            for row_idx, execution in enumerate(executions, start=2):
                flat_exec = self.exporter._flatten_dict(execution)
                for col_idx, header in enumerate(headers, start=1):
                    value = flat_exec.get(header, "")
                    ws_executions.cell(row=row_idx, column=col_idx, value=str(value))
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _generate_hitl_excel_report(
        self,
        records: List[Dict[str, Any]],
        analytics: Dict[str, Any]
    ) -> bytes:
        """Generate HITL Excel report with analytics"""
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        row = 1
        for key, value in analytics.items():
            ws_summary.cell(row=row, column=1, value=key)
            ws_summary.cell(row=row, column=2, value=value)
            row += 1
        
        # Records sheet
        ws_records = wb.create_sheet("HITL Records")
        
        if records:
            headers = list(self.exporter._flatten_dict(records[0]).keys())
            for col_idx, header in enumerate(headers, start=1):
                ws_records.cell(row=1, column=col_idx, value=header)
            
            for row_idx, record in enumerate(records, start=2):
                flat_record = self.exporter._flatten_dict(record)
                for col_idx, header in enumerate(headers, start=1):
                    value = flat_record.get(header, "")
                    ws_records.cell(row=row_idx, column=col_idx, value=str(value))
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()


# Global instances
_exporter = DataExporter()
_report_generator = ReportGenerator(_exporter)


def get_exporter() -> DataExporter:
    """Get global data exporter instance"""
    return _exporter


def get_report_generator() -> ReportGenerator:
    """Get global report generator instance"""
    return _report_generator
